import openpyxl

# Creating and saving excel documents
wb = openpyxl.Workbook()                # Create a blank workbook
wb.sheetnames                           # It starts with one sheet
sheet = wb.active
sheet.title
sheet.title = 'Spam Bacon Eggs Sheet'   # Change sheet title
wb.sheetnames
wb.save('example_copy.xlsx')            # Save the workbook

# Creating and Removing Sheets
wb.create_sheet()                               # Add a new sheet
wb.sheetnames
wb.create_sheet(index=0, title='First Sheet')   # Create new sheet at index 0
wb.sheetnames
wb.create_sheet(index=2, title='Middle Sheet')
wb.sheetnames
del wb['Middle Sheet']                          # Delete sheet

# Writing values to cells
sheet['A1'] = 'Hello, world!' # Edit the cell's value
sheet['A1'].value
