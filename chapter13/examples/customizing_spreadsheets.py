import openpyxl
from openpyxl.styles import Font

# Setting the Font Style
wb = openpyxl.Workbook()
sheet = wb.active
italic_24_font = Font(size=24, italic=True) # Create a font
sheet['A1'].font = italic_24_font # Apply the font to A1
sheet['A1'] = 'Hello, world!'
wb.save('styles.xlsx')

font_obj1 = Font(name='Liberation Serif', bold=True)
sheet['B3'].font = font_obj1
sheet['B3'] = 'Bold Liveration Serif'

font_obj2 = Font(size=24, bold=True)
sheet['C5'].font = font_obj2
sheet['C5'] = '24 pt Italic'

wb.save('styles.xlsx')

# An excel formula is set just like any other text value in a cell
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # Set the formula
wb.save('write_formula.xlsx')

# Seting row height and column width
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# Merging and unmerging cells
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # Merge all these cells
sheet['A1'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5') # Merge these two cells
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')
sheet.unmerge_cells('A1:D3') # Split these cells
sheet.unmerge_cells('C5:D5') 
wb.save('merged.xlsx')

# Freezing panes
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2' # Freeze the rows above A2
wb.save('freeze_example.xlsx')
