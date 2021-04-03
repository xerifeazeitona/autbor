"""
Excel-to-CSV Converter

Using the openpyxl module from Chapter 13, write a program that reads
all the Excel files in the current working directory and outputs them
as CSV files.

A single Excel file might contain multiple sheets; you’ll have to create
one CSV file per sheet. The filenames of the CSV files should be
<excel filename>_<sheet title>.csv, where <excel filename> is the
filename of the Excel file without the file extension (for example,
'spam_data', not 'spam_data.xlsx') and <sheet title> is the string from
the Worksheet object’s title variable.
"""

import os
import csv
import openpyxl

os.chdir(os.path.join(os.getcwd(), 'sample_data'))

for filename in os.listdir():
    if not filename.endswith('.xlsx'):
        continue # skip non-xlsx files
    print(f'Converting {filename}...')
    # Open the spreadsheet
    wb = openpyxl.load_workbook(filename)
    # Create a CSV file for each sheet
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]

        # create csv filename from excel filename and sheet title
        csv_filename = f'{os.path.splitext(filename)[0]}_{sheetname}.csv'

        # create csvwriter object
        output_file = open(csv_filename, 'w', newline='')
        output_writer = csv.writer(output_file)

        # Loop through every row in the sheet
        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            # Loop through every column in the sheet
            for col_num in range(1, sheet.max_column + 1):
                row_data.append(sheet.cell(row=row_num, column=col_num).value)
            output_writer.writerow(row_data)

        # Save csv file
        output_file.close()
print('...done!')
