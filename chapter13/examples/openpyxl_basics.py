import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Opening excel documents with openpyxl
wb = openpyxl.load_workbook('example.xlsx')

# Getting sheets from the Workbook
wb.sheetnames               # The workbook's sheets' names
sheet = wb['Sheet1']        # Get a sheet from the workbook
sheet.title                 # Get the sheet's title as a string
another_sheet = wb.active   # Get the active sheet

# Getting cells from the Sheets
sheet['A1']         # Get a cell from the sheet
sheet['A1'].value   # Get the value from the cell
c = sheet['B1']     # Get another cell from the sheet
c.value
# Get the row, column and value from the cell
print(f'Row {c.row}, Column {c.column}, is {c.value}')
print(f'Cell {c.coordinate} is {c.value}')
# You can also get a cell using the sheet’s cell() method
sheet.cell(row=1, column=2).value
for i in range(1, 8, 2): # Go through every other row
    print(i, sheet.cell(row=i, column=2).value)
# Get the size of the sheet with the objects max_row and max_column
sheet.max_row       # Get the highest row number
sheet.max_column    # Get the highest column number

# Converting between column letters and numbers
get_column_letter(1)            # Translate cloumn 1 to a letter
get_column_letter(2)
get_column_letter(27)
get_column_letter(900)
get_column_letter(sheet.max_column)
column_index_from_string('A')   # Get A's number
column_index_from_string('AA')

# Getting rows and columns from the sheets
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3
for row_of_cell_objects in sheet['A1':'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print('--- END OF ROW ---')

list(sheet.columns)[1] # Get the second column's cells
for cell_obj in list(sheet.columns)[1]:
    print(cell_obj.value)
